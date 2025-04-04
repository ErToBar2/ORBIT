import numpy as np
import random
from scipy.interpolate import interp1d
from PySide2.QtWidgets import QMessageBox
from PySide2.QtWidgets import QInputDialog
##### compute anlges between pillars and the bridge


import os
import glob
import shutil
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pyproj import Transformer

def write_data_to_excel(excel_file_path, trajectory_3D_list, pillars_3D_list):
    # Prepare the transformer to convert WGS84 to Lambert (Belgian Lambert 72 as an example)
    transformer = Transformer.from_crs("epsg:4326", "epsg:31370", always_xy=True)

    # Prepare DataFrame
    columns = ['nr', 'x', 'y', 'z', 'Element']
    data = []

    # Process trajectory data and transform coordinates
    for index, (x, y, z) in enumerate(trajectory_3D_list, start=1):
        x_lambert, y_lambert = transformer.transform(y, x)
        nr = f'36{index:02}'
        data.append([nr, x_lambert, y_lambert, z, 'BridgeTrajectory'])

    # Process pillars data and transform coordinates
    for index, (x, y, z) in enumerate(pillars_3D_list, start=1):
        x_lambert, y_lambert = transformer.transform(y, x)
        nr_right = f'13{index:02}' if index % 2 != 0 else f'23{index - 1:02}'  # Adjusted for right (odd index) and left (even index)
        element_type = 'Pier'
        data.append([nr_right, x_lambert, y_lambert, z, element_type])

    df = pd.DataFrame(data, columns=columns)

    # Load or create the workbook
    try:
        book = load_workbook(excel_file_path)
    except FileNotFoundError:
        book = Workbook()
        book.save(excel_file_path)
        book = load_workbook(excel_file_path)

    # Manage '01_data_from_map' sheet
    if '01_data_from_map' in book.sheetnames:
        del book['01_data_from_map']  # Delete the sheet if it exists
    ws = book.create_sheet('01_data_from_map')  # Create a new sheet
    
    # Write DataFrame to the new sheet using openpyxl
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    book.save(excel_file_path)
    book.close()

    print("Data successfully written to:", excel_file_path)

###################################################################################################################

import numpy as np

class TrajectoryAnalyzer:
    def __init__(self, trajectory, pillar_centers):
        self.trajectory = trajectory
        self.pillar_centers = pillar_centers

    @staticmethod
    def calculate_xy_distance(point1, point2):
        """Calculate the Euclidean distance between two points in the XY plane."""
        point1 = np.array(point1[:2])  # Consider only X and Y coordinates
        point2 = np.array(point2[:2])  # Consider only X and Y coordinates
        return np.linalg.norm(point1 - point2)

    def calculate_total_length_xy(self):
        """Calculate the total length of the trajectory in the XY plane."""
        total_length = 0
        for i in range(1, len(self.trajectory)):
            total_length += self.calculate_xy_distance(self.trajectory[i - 1], self.trajectory[i])
        return total_length

    def calculate_distances_between_pillars(self):
        """Calculate distances between pillars and from the start/end of the bridge."""
        distances = []
        if self.pillar_centers:
            distances.append(self.calculate_xy_distance(self.trajectory[0], self.pillar_centers[0]))

        for i in range(1, len(self.pillar_centers)):
            distances.append(self.calculate_xy_distance(self.pillar_centers[i - 1], self.pillar_centers[i]))

        distances.append(self.calculate_xy_distance(self.pillar_centers[-1], self.trajectory[-1]))
        return distances

    def print_distances(self):
        """Prints formatted distances from start to first pillar, between pillars, and last pillar to end."""
        distances = self.calculate_distances_between_pillars()
        if distances:
            print("Distance from start of the bridge to the first pillar: {:.2f} meters".format(distances[0]))
            for i in range(1, len(distances) - 1):
                print("Distance from pillar {} to pillar {}: {:.2f} meters".format(i, i + 1, distances[i]))
            print("Distance from the last pillar to the end of the bridge: {:.2f} meters".format(distances[-1]))

        return distances

#################################################


def setup_GUI_paths(bridge_name, base_directory):
    """ Sets up the necessary directories and prepares files for flight path planning associated with a specific bridge.

    This function creates specific directories for visualizations, input data, and flight routes based on the bridge's name.
    It then copies relevant files from a global input directory to a bridge-specific input directory and checks for the
    existence of a cross-section image and an Excel file relevant to the bridge. It also prints the names of all copied files
    and provides status updates on the existence of the cross-section image and Excel file.

    Parameters:
        bridge_name (str): The name of the bridge, which is used to tailor directory names and identify specific files.
        base_directory (str): formulates where to find the primary input files  = os.path.join(base_directory, "01_Input", f"*{bridge_name}*.*")

    Returns:
        tuple: Contains paths to the primary input directory, input directory, visualization directory, flight route directory,
            path to the cross-section image (if found), and path to the Excel file (if found).

    Each directory and file path returned is specific to the provided bridge name, facilitating """

    # Define directories based on bridge name
    visuals_directory = os.path.join(base_directory, bridge_name, "02_Visualization")
    input_directory = os.path.join(base_directory, bridge_name, "01_Input")
    flightroute_directory = os.path.join(base_directory, bridge_name, "03_Flightroutes")

    # Create directories if they do not exist
    os.makedirs(visuals_directory, exist_ok=True)
    os.makedirs(input_directory, exist_ok=True)
    os.makedirs(flightroute_directory, exist_ok=True)

    # Copy relevant files from the global input folder
    primaryInputDir  = os.path.join(base_directory, "01_Input", f"*{bridge_name}*.*")
    for file in glob.glob(primaryInputDir):
        shutil.copy(file, input_directory)
        print(os.path.basename(file))

    # Identify the cross-section image
    crosssection_dir = None
    image_extensions = ['.png', '.jpg']
    for ext in image_extensions:
        crosssection_path = os.path.join(input_directory, f"{bridge_name}_crosssection_edit{ext}")
        if os.path.exists(crosssection_path):
            crosssection_dir = crosssection_path
            break
        
    # Identify the Excel file
    excel_file_path = None  # Use one variable to hold the final path
    excel_extensions = ['.xlsx']
    for ext in excel_extensions:
        possible_excel_file_path = os.path.join(input_directory, f"{bridge_name}{ext}")
        if os.path.exists(possible_excel_file_path):
            excel_file_path = possible_excel_file_path
            break

    # Print results
    if crosssection_dir:
        print(f"Cross-section image found: {crosssection_dir}")
    else:
        print("Cross-section image not found.")   
    
    # Print results regarding the Excel file
    if excel_file_path:
        print(f"Excel found: {excel_file_path}")
    else:
        print("No Excel found. Creating a new Excel file.")
        excel_file_path = os.path.join(input_directory, f"{bridge_name}.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '00_Input'
        columns = ['nr.', 'X', 'Y', 'Z', 'Element']
        sheet.append(columns)  # Adding the column headers
        workbook.save(filename=excel_file_path)
        print(f"Excel created: {excel_file_path}")
    
    return primaryInputDir, input_directory, visuals_directory,flightroute_directory, crosssection_dir, excel_file_path


############################################################################################################################

import pandas as pd
from openpyxl import load_workbook, Workbook
from pyproj import Transformer
import os
import xml.etree.ElementTree as ET

def robust_convert_to_float(value):
    try:
        # Try to directly convert to float if possible
        return float(value)
    except ValueError:
        try:
            # Remove spaces and try to convert again
            cleaned_value = str(value).replace(' ', '')  # Remove any spaces from the string
            return float(cleaned_value)  # Convert the cleaned string to float
        except ValueError:
            # If conversion still fails, return NaN
            return float('nan')
        
def convert_columns(df, column_names):
    """Convert specified columns in the DataFrame to numeric types."""
    for column in column_names:
        df[column] = df[column].apply(robust_convert_to_float)
    return df

def load_and_transform_excel(input_path):
    """Loads the Excel file and transforms coordinates if necessary, handling specific conditions."""
    df = pd.read_excel(input_path, sheet_name='00_Input')
    # Convert 'nr.' column to string
    df['nr.'] = df['nr.'].astype(str)

    # Identify columns that exist in the dataframe for conversion
    columns_to_convert = ['X', 'Y', 'Z']  # Always present columns
    if 'X_Google' in df.columns and 'Y_Google' in df.columns:
        columns_to_convert += ['X_Google', 'Y_Google']
    
    # Apply the robust conversion function to existing columns
    df = convert_columns(df, columns_to_convert)

    # Check for existence of 'X_Google' and 'Y_Google' and condition of 'X' and 'Y' columns
    if 'X_Google' in df.columns and 'Y_Google' in df.columns:
        if ('X' not in df.columns or df['X'].isna().all()) and ('Y' not in df.columns or df['Y'].isna().all()):
            # Perform the conversion only if both 'X' and 'Y' columns are either missing or all NaN
            df['X'], df['Y'] = zip(*df.apply(lambda row: convert_wgs84_to_lambert(row['X_Google'], row['Y_Google']), axis=1))
            df.drop(['X_Google', 'Y_Google'], axis=1, inplace=True)

    return df


def convert_wgs84_to_lambert(lat, lon):
    """Converts coordinates from WGS84 to Lambert projection."""
    transformer = Transformer.from_crs("EPSG:4326", "EPSG:31370", always_xy=True)
    return transformer.transform(lon, lat)

def translate_element(nr):
    """Maps numeric codes to structural elements."""
    nr = str(nr)  # Ensure nr is treated as a string
    mapping = {
        '1': 'Abutment',
        '2': 'Superstructure',
        '3': 'Pier',
        '4': 'Bottom of Beam/Construction',
        '5': 'Wall',
        '6': 'BridgeTrajectory'
    }
    return mapping.get(nr[1], 'Unknown')

def generate_middle_trajectory(df):
    """Generates a middle trajectory for superstructure elements by averaging right and left entries, including X, Y, and Z values."""
    super_df = df[df['nr.'].str[1] == '2']  # Filter superstructure elements
    grouped = super_df.groupby(super_df['nr.'].str[2:])  # Group by the last part of 'nr.'

    middle_entries = []
    for suffix, group in grouped:
        if len(group) == 2:  # Ensure there are both right and left entries
            middle_x = group['X'].mean()
            middle_y = group['Y'].mean()
            middle_z = group['Z'].mean()
            middle_nr = '36' + suffix  # '3' for middle, '6' for BridgeTrajectory, with existing suffix
            middle_entries.append({
                'nr.': middle_nr,
                'X': middle_x,
                'Y': middle_y,
                'Z': middle_z,
                'Element': 'BridgeTrajectory'
            })

    if middle_entries:
        middle_df = pd.DataFrame(middle_entries)
        df = pd.concat([df, middle_df], ignore_index=True)

    return df

def adjust_z_values(df, trajectory_heights):
    """Adjusts Z values based on availability and specific rules, only filling missing values."""
    update_trajectory_data_necessary = False
    height_index = 0  # Initialize an index to keep track of which height to use next

    for i, row in df.iterrows():
        if pd.isna(row['Z']):  # Check if Z value is NaN
            if row['Element'] == 'BridgeTrajectory':
                # Only update if there are available heights left in the list
                if height_index < len(trajectory_heights):
                    df.at[i, 'Z'] = trajectory_heights[height_index]
                    height_index += 1  # Increment to use the next height for the next missing Z
                    update_trajectory_data_necessary = True  
            elif row['Element'] in ['Pier', 'Abutment']:
                # Set Z to 0 for 'Pier' and 'Abutment' if Z is NaN
                df.at[i, 'Z'] = 0

    return df, update_trajectory_data_necessary





def parse_kml(kml_file_path):
    # Initialize tree and namespace
    tree = ET.parse(kml_file_path)
    root = tree.getroot()
    ns = {'kml': 'http://www.opengis.net/kml/2.2'}
    
    # Setup transformation from WGS84 to Lambert 72
    transformer = Transformer.from_crs("epsg:4326", "epsg:31370", always_xy=True)

    # Initialize data holders
    kml_data = {'Right': [], 'Middle': [], 'Left': []}

    # Parse placemarks and extract coordinates
    for placemark in root.findall('.//kml:Placemark', ns):
        name = placemark.find('kml:name', ns).text.strip()
        coordinates = placemark.find('.//kml:coordinates', ns).text.strip()

        # Process each coordinate set
        coords = []
        for coord_set in coordinates.split():
            lon, lat, alt = map(float, coord_set.split(','))
            x_lambert, y_lambert = transformer.transform(lon, lat)
            coords.append((x_lambert, y_lambert, alt))
        
        if name in kml_data:
            kml_data[name] = coords

    return kml_data['Right'], kml_data['Middle'], kml_data['Left']

def update_df_with_kml(df, right_data, middle_data, left_data):
    # Decide which entries to delete
    if right_data or left_data:
        if right_data:
            df = df[~df['nr.'].str.startswith('12')]
        if left_data:
            df = df[~df['nr.'].str.startswith('22')]
        if middle_data or (right_data and left_data):
            df = df[~df['nr.'].str.startswith('36')]

    # Prepare to insert new data
    new_entries = []
    for idx, (x, y, z) in enumerate(right_data, start=1):
        new_entries.append({'nr.': f'12{idx:02d}', 'X': x, 'Y': y, 'Z': z, 'Element': 'Superstructure'})
    for idx, (x, y, z) in enumerate(left_data, start=1):
        new_entries.append({'nr.': f'22{idx:02d}', 'X': x, 'Y': y, 'Z': z, 'Element': 'Superstructure'})
    for idx, (x, y, z) in enumerate(middle_data, start=1):
        new_entries.append({'nr.': f'36{idx:02d}', 'X': x, 'Y': y, 'Z': z, 'Element': 'BridgeTrajectory'})

    # Add new entries
    if new_entries:
        new_df = pd.DataFrame(new_entries)
        df = pd.concat([df, new_df], ignore_index=True)

    # Generate middle trajectory if applicable
    if right_data and left_data:
        df = generate_middle_trajectory(df)

    return df



def extract_special_points(df):
    """Extracts special points for bridge trajectories and pillars and returns them as lists."""
    # Filter the DataFrame for Bridge Trajectory elements and extract relevant columns
    trajectory_df = df[df['Element'] == 'BridgeTrajectory']
    trajectory_3d_list = trajectory_df[['X', 'Y', 'Z']].values.tolist()

    # Filter for Pillar elements and sort by last two digits, then by the full 'nr.' column
    pillars_df = df[df['Element'] == 'Pier'].copy()
    
    # Create a new column for sorting by the last two digits
    pillars_df['suffix'] = pillars_df['nr.'].apply(lambda x: x[-2:])
    # Now sort by suffix first, and then by 'nr.' to ensure correct pairing
    sorted_pillars_df = pillars_df.sort_values(by=['suffix', 'nr.'])
    pillars_3d_list = sorted_pillars_df[['X', 'Y', 'Z']].values.tolist()

    # Clean up the DataFrame by dropping the 'suffix' column if no longer needed
    pillars_df.drop(columns='suffix', inplace=True)

    return trajectory_3d_list, pillars_3d_list

def save_to_excel(df, excel_file_path):
    """Saves the DataFrame to an Excel file, overwriting the existing '00_Input' sheet."""
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='00_Input', index=False)

    
def dataloader_excel(excel_file_path, input_directory, trajectory_heights, bridge_name):
    """Loads data from Excel, updates it with optional KML data, and writes back to the same Excel file."""
    df = load_and_transform_excel(excel_file_path)
    df['Element'] = df['nr.'].apply(translate_element)
    df = generate_middle_trajectory(df)
    df, update_trajectory_data_necessary  = adjust_z_values(df, trajectory_heights)
    
    kml_file_path = os.path.join(input_directory, f"{bridge_name}.kml")
    if os.path.exists(kml_file_path):
        print("Updating excel using KML file found:", kml_file_path)
        right_data, middle_data, left_data = parse_kml(kml_file_path)
        df = update_df_with_kml(df, right_data, middle_data, left_data)
    
    trajectory_3d_list, pillars_3D_list = extract_special_points(df)
    # Debugging or further processing with the lists
    print("Trajectory Points:", trajectory_3d_list)
    print("Pillar Points:", pillars_3D_list)

    save_to_excel(df, excel_file_path)

    return trajectory_3d_list, pillars_3D_list, update_trajectory_data_necessary



def calculate_distance(point1, point2):
    """Calculate Euclidean distance between two points (supports 2D or 3D)."""
    return np.linalg.norm(np.array(point1) - np.array(point2))


def calculate_nearest_point_on_curve(curve, external_point):
    """Find the nearest point on the curve to an external point and its index."""
    min_distance = float('inf')
    nearest_point_index = None
    for i, point in enumerate(curve):
        distance = calculate_distance(point, external_point)
        if distance < min_distance:
            min_distance = distance
            nearest_point_index = i
    return nearest_point_index

def calculate_cumulative_distance_to_index(curve, index):
    """Calculate cumulative distance up to a specified index in the curve."""
    total_distance = 0
    for i in range(1, index + 1):
        total_distance += calculate_distance(curve[i - 1][0], curve[i][0])
    return total_distance

def calculate_vector(point1, point2):
    """Calculate a 2D vector from two 3D points by ignoring the z-coordinate."""
    return np.array([point2[0] - point1[0], point2[1] - point1[1]])

def magnitude(vector):
    """Calculate the magnitude of a 2D vector."""
    return np.sqrt(vector[0]**2 + vector[1]**2)


def calculate_perpendicular_angle(vector1, vector2):
    """Calculate the angle between two vectors, considering perpendicular as 0 degrees."""
    dot_product = np.dot(vector1, vector2)
    angle_cos = dot_product / (magnitude(vector1) * magnitude(vector2))
    angle_rad = np.arccos(np.clip(angle_cos, -1.0, 1.0))  # Clipping for numerical stability
    angle_deg = np.degrees(angle_rad)
    # To consider perpendicular as 90 degrees, calculate the deviation from 0 degrees
    return 90 - abs(angle_deg)

def compute_local_bridge_vector(trajectory, index):
    """Compute the local bridge vector at a given index, using only x and y coordinates."""
    prev_index = max(0, index - 1)
    next_index = min(len(trajectory) - 1, index + 1)
    # Ensuring only x and y coordinates are used
    return calculate_vector(trajectory[prev_index][:2], trajectory[next_index][:2])


################### compute points in angle and distance from trajecotrty index points. 
def calculate_point_at_distance_along_curve(curve, distance_along_curve, normals, perpendicular_distance=None):
    """
    Calculate a point at a certain distance along the curve and optionally
    calculate a perpendicular point at a specified distance.
    
    Parameters:
    - curve: List of points defining the curve (3D points).
    - distance_along_curve: Distance from the starting point along the curve.
    - normals: List of normal vectors at each point in the curve.
    - perpendicular_distance: Distance to calculate the point perpendicular to the curve (optional).
    
    Returns:
    - point_on_curve: The point on the curve at the specified distance.
    - perpendicular_point: The point perpendicular to the curve at the specified distance (if requested).
    """
    accumulated_distance = 0
    for i in range(1, len(curve)):
        prev_point = curve[i - 1]
        point = curve[i]
        segment_distance = calculate_distance(prev_point, point)
        if accumulated_distance + segment_distance >= distance_along_curve:
            # Calculate point on curve
            ratio = (distance_along_curve - accumulated_distance) / segment_distance
            point_on_curve = prev_point + ratio * (point - prev_point)

            # Calculate perpendicular point if required
            if perpendicular_distance is not None:
                normal = normals[i - 1]  # Use the normal at the previous point or interpolate
                perpendicular_point = point_on_curve + perpendicular_distance * normal
                return point_on_curve, perpendicular_point
            return point_on_curve, None

        accumulated_distance += segment_distance

    raise ValueError("The specified distance exceeds the length of the curve.")

def compute_base_points(trajectory, distances_pillars, thresholds_zones, num_points):
    sections_base_points = []
    sections_indices = []
    start_distance = 0

    for section_index, distance in enumerate(distances_pillars):
        end_distance = start_distance + distance
        section_start_distance = start_distance + (thresholds_zones[section_index][0] if section_index < len(thresholds_zones) else 0)
        section_end_distance = end_distance - (thresholds_zones[section_index][1] if section_index < len(thresholds_zones) else 0)

        # Calculate the total length of the current section
        section_length = section_end_distance - section_start_distance
        if section_length <= 0 or num_points[section_index] <= 0:
            continue

        # Calculate intervals between base points within the section
        if num_points[section_index] == 1:
            interval = section_length  # Ensures one point is placed at the start if only one is requested
        else:
            interval = section_length / (num_points[section_index] - 1)

        current_distance = 0
        accumulated_distance = 0
        section_points = []
        section_point_indices = []
        found_base_points = 0

        for i in range(1, len(trajectory)):
            if found_base_points >= num_points[section_index]:
                break

            prev_point = np.array(trajectory[i - 1])
            current_point = np.array(trajectory[i])
            segment_length = np.linalg.norm(current_point - prev_point)

            while accumulated_distance + segment_length >= section_start_distance + found_base_points * interval:
                if found_base_points >= num_points[section_index]:
                    break

                # Distance from the start of the segment to the required base point
                distance_into_segment = (section_start_distance + found_base_points * interval) - accumulated_distance

                # Interpolate to find the base point
                ratio = distance_into_segment / segment_length
                base_point = prev_point + ratio * (current_point - prev_point)
                section_points.append(base_point.tolist())
                section_point_indices.append(i - 1)  # Use i-1 to denote the segment starting index
                found_base_points += 1

            accumulated_distance += segment_length

        sections_base_points.append(section_points)
        sections_indices.append(section_point_indices)
        start_distance = end_distance  # Update the start distance for the next section

    return sections_base_points, sections_indices


def adjust_zone_angles(pillar_angles, num_zones):
    if len(pillar_angles) >= num_zones:
        return pillar_angles[:num_zones]
    
    # Create a new list of angles
    adjusted_angles = [pillar_angles[0]]  # Start with the first angle

    # Interpolate the middle angles
    for i in range(1, num_zones - 1):
        if i < len(pillar_angles) - 1:
            # Calculate mean of this and the next original angle
            mean_angle = (pillar_angles[i] + pillar_angles[i - 1]) / 2
            adjusted_angles.append(mean_angle)
        else:
            # Use the last interpolated angle again if out of original angles
            adjusted_angles.append((adjusted_angles[-1] + pillar_angles[-1]) / 2)

    # Append the last original angle
    adjusted_angles.append(pillar_angles[-1])

    return adjusted_angles


# For the underdeck flgiht route:
def adjust_height_offsets_with_quadratic(height_offsets, num_points):
    adjusted_heights = []
    for offsets, num in zip(height_offsets, num_points):
        if len(offsets) == num:
            adjusted_heights.append(offsets)
        else:
            # Define x coordinates for existing offsets and for the desired number of points
            x_existing = np.linspace(0, 1, len(offsets))
            x_new = np.linspace(0, 1, num)
            
            # If only one offset, we can't do much, just repeat it
            if len(offsets) == 1:
                interpolated = [offsets[0]] * num
            else:
                # Create a quadratic interpolator
                f = interp1d(x_existing, offsets, kind='quadratic', fill_value="extrapolate")
                interpolated = f(x_new).tolist()

            adjusted_heights.append(interpolated)
    return adjusted_heights


def interpolate_trajectory_height(heights, target_length):
    """
    Interpolates or truncates the heights list to match the target_length.
    """
    if len(heights) == target_length:
        return heights
    elif len(heights) < target_length:
        x_original = list(range(len(heights)))
        x_target = list(range(target_length))
        interpolated = np.interp(x_target, x_original, heights)
        return interpolated.tolist()
    else:
        return heights[:target_length]

def calculate_adjusted_point(point, perpendicular_distance, normal, height_offset, angle_degrees):
    """Calculate a point on an angle to the curve, rotate it by a given angle around the Z axis in degrees, and adjust its height."""
    #print(f"Original angle in degrees: {angle_degrees}")  # Debug print

    # Convert angle from degrees to radians
    angle_radians = np.deg2rad(angle_degrees)
    #print(f"Converted angle in radians: {angle_radians}")  # Debug print

    # Normalize the normal vector
    normalized_normal = normal / np.linalg.norm(normal)
    #print(f"Normalized normal vector: {normalized_normal}")  # Debug print

    # Calculate rotation matrix for Z-axis rotation
    rotation_matrix = np.array([
        [np.cos(angle_radians), -np.sin(angle_radians), 0],
        [np.sin(angle_radians), np.cos(angle_radians), 0],
        [0, 0, 1]
    ])

    # Rotate the normalized normal vector
    rotated_normal = np.dot(rotation_matrix, normalized_normal)
    #print(f"Rotated normal vector: {rotated_normal}")  # Debug print

    # Calculate the adjusted point with the rotated vector
    adjusted_point = point + perpendicular_distance * rotated_normal
    
    # Adjust the Z component of the point
    adjusted_point[2] -=  height_offset  #
    #print(f"Adjusted point: {adjusted_point}")  # Debug print

    return adjusted_point

def compute_points_with_horizontal_offset(base_points, normals, horizontal_offsets, height_offsets, angles):
    offset_points = []
    for section_index, points in enumerate(base_points):
        section_points = []
        for i, base_point in enumerate(points):
            normal = normals[i % len(normals)]
            angle = angles[section_index]
            height_offset = height_offsets[section_index][i % len(height_offsets[section_index])]

            # Calculate right and left points
            point_right = calculate_adjusted_point(np.array(base_point), horizontal_offsets[section_index], normal, height_offset, angle)
            point_left = calculate_adjusted_point(np.array(base_point), -horizontal_offsets[section_index], normal, height_offset, angle)

            # Convert numpy arrays to lists
            section_points.append(point_right.tolist())
            section_points.append(point_left.tolist())

        offset_points.append(section_points)
    return offset_points




def compute_safe_flythrough_path(num_points, offset_points_underdeck, parent=None):
 
    # Identify the middle section index
    total_sections = len(num_points)
    middle_section_index = total_sections // 2

    # If the number of sections is even, ask the user to select which middle section to use
    if total_sections % 2 == 0:
        msg = QMessageBox(parent)
        msg.setIcon(QMessageBox.Information)
        msg.setText("Even number of sections detected.")
        msg.setInformativeText("Please select the section index you'd like to process:")
        msg.setWindowTitle("Section Selection")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()

        # Let the user input the section index
        section, ok = QInputDialog.getInt(parent, "Select Section", "Enter section index (1-based):", middle_section_index, 1, total_sections, 1)
        if ok:
            middle_section_index = section - 1  # Convert 1-based index to 0-based
        else:
            # User cancelled or closed the dialog, handle appropriately
            return []
    underdeck_safe_flythrough=[]
    middle_section_points = offset_points_underdeck[middle_section_index]
    # Retrieve the "middle_section_points" section's offset points (offset_points_underdeck[middle_section_index] ) and devide by 2
    num_passes = len(middle_section_points) //2
    print(f"num_passes: {num_passes}")
    # Determine the most central pair(s) of points
    if num_passes % 2 == 1:
        # Odd number of passes, one central pair
        central_index = num_passes // 2
        underdeck_safe_flythrough.append(middle_section_points[2 * central_index])
        underdeck_safe_flythrough.append(middle_section_points[2 * central_index + 1])
    else:

        # Even number of passes, interpolate between two central pairs
        first_central_index = num_passes // 2 - 1
        second_central_index = num_passes // 2

        # Get the points for the two central pairs
        first_pair_right = middle_section_points[2 * first_central_index]
        first_pair_left = middle_section_points[2 * first_central_index + 1]
        second_pair_right = middle_section_points[2 * second_central_index]
        second_pair_left = middle_section_points[2 * second_central_index + 1]

        # Interpolate the right points and convert to tuple
        interpolated_right = tuple((first_pair_right[i] + second_pair_right[i]) / 2 for i in range(3))
        # Interpolate the left points and convert to tuple
        interpolated_left = tuple((first_pair_left[i] + second_pair_left[i]) / 2 for i in range(3))

        # Append interpolated points separately as 3D tuples
        underdeck_safe_flythrough.append(interpolated_right)
        underdeck_safe_flythrough.append(interpolated_left)
    return underdeck_safe_flythrough

#### UNDERDECK AXIAL 
def compute_girder_offsets(bridge_width, n_girders):
    if n_girders % 2 == 1:
        # Uneven number of girders
        half_girders = (n_girders - 1) // 2
        spacing = bridge_width / (n_girders - 1)
        offsets = [-half_girders * spacing + i * spacing for i in range(n_girders)]
    else:
        # Even number of girders
        half_width = bridge_width / 2
        spacing = bridge_width / n_girders
        offsets = [-half_width + spacing/2 + i * spacing for i in range(n_girders)]
    return offsets


def points_underdeck_flight_route_axial(base_points, normals, bridge_width, n_girders, height_offsets, angles,offset_points_underdeck,connection_height_value):
    # Compute girder offsets sorted from nearest to farthest from the centerline of the bridge
    girder_offsets = sorted(compute_girder_offsets(bridge_width, n_girders), reverse=True)
    zigzag_points = []
    
    for section_index, points in enumerate(base_points):
        section_points = []
        
        # Start the section by appending the very first offset point
        first_offset_point = offset_points_underdeck[section_index][0]
        section_points.extend([first_offset_point, [first_offset_point[0], first_offset_point[1], first_offset_point[2] + connection_height_value], first_offset_point])

        # Process each offset for the current section
        for offset_index, offset in enumerate(girder_offsets):
            forward_points = []

            for i, base_point in enumerate(points):
                normal = normals[i % len(normals)]
                angle = angles[section_index]
                height_offset = height_offsets[section_index][i % len(height_offsets[section_index])]
                adjusted_point = calculate_adjusted_point(np.array(base_point), offset, normal, height_offset, angle)
                forward_points.append(adjusted_point.tolist())
            
            # after reaching the end of the first offset, append the connection flight
            if offset_index == 0:
                last_right_point_index = len(offset_points_underdeck[section_index]) - 2
                connection_offset_point = offset_points_underdeck[section_index][last_right_point_index]
                forward_points.extend([connection_offset_point, [connection_offset_point[0], connection_offset_point[1], connection_offset_point[2] + connection_height_value]])
            # reaching the last offset, append the third alast connection flight
            if offset_index == len(girder_offsets) - 1 :
                last_left_point_index = len(offset_points_underdeck[section_index]) - 1
                connection_offset_point = offset_points_underdeck[section_index][last_left_point_index]
                forward_points.extend([connection_offset_point, [connection_offset_point[0], connection_offset_point[1], connection_offset_point[2] + connection_height_value]])
                
            # Append the forward points to the section points
            section_points.extend(forward_points)
            backward_points = forward_points[::-1]
            section_points.extend(backward_points)


        # Finish the section by appending the very last offset point
        last_offset_point = offset_points_underdeck[section_index][1]
        section_points.extend([last_offset_point, [last_offset_point[0], last_offset_point[1], last_offset_point[2] + connection_height_value], last_offset_point])

        zigzag_points.append(section_points)

    return zigzag_points




    #         # Backward_points is the reversed order of forward_points:
    #         backward_points = forward_points[::-1]

    #         # Append forward or backward points to section points
    #         if append_forward:
    #             section_points.extend(forward_points)
    #             append_forward = False  # Next offset will append backward point
    #         else:
    #             section_points.extend(backward_points)
    #             append_forward = True  # Next offset will append forward points

    #     # Finish the section by appending the very last offset point
    #     last_offset_point = offset_points_underdeck[section_index][-1]
    #     section_points.extend([last_offset_point, [last_offset_point[0], last_offset_point[1], last_offset_point[2] + connection_height_value], last_offset_point])

    #     zigzag_points.append(section_points)

    # return zigzag_points


###### Visualization 

def show_start(flight_route):
    # Check if the flight_route is not empty
    if not flight_route:
        print("Error: The flight route list is empty.")
        return flight_route  # Return the unmodified list or handle as needed
    
    # Check if the first point has at least three elements (x, y, z)
    if len(flight_route[0]) < 3:
        print("Error: The flight route points must include at least x, y, and z coordinates.")
        return flight_route  # Return the unmodified list or handle as needed
    
    # Make a copy of the flight route to avoid modifying the original list
    modified_flight_route = flight_route.copy()
    
    # Make a copy of the first point
    new_start_point = modified_flight_route[0][:]  # Use slicing to copy the list
    
    # Set the z-coordinate to 0
    new_start_point[2] = 0

    # Insert this new point at the beginning of the list
    modified_flight_route.insert(0, new_start_point)
    
    # Return the modified flight route list
    return modified_flight_route

def clean_trajectory(trajectory_list):
    # Initialize a new list to store the cleaned trajectory points
    cleaned_list = []
    
    # Loop through each element in the trajectory list
    for element in trajectory_list:
        # Check if the element is a list
        if isinstance(element, list):
            # Check if the first element of this list is also a list, suggesting nested lists
            if element and isinstance(element[0], list):
                # Recursively clean the nested list
                cleaned_list.append(clean_trajectory(element))
            else:
                # It's a single point; ensure the point has at least three coordinates
                if len(element) >= 3:
                    # Extract the first three elements (coordinates) and ignore any extra elements (tags)
                    cleaned_point = element[:3]
                    # Append the cleaned point to the cleaned list
                    cleaned_list.append(cleaned_point)
    
    # Return the cleaned trajectory list
    return cleaned_list

def write_obj_with_lines(file_path, vertices):
    """
    Write a 3D object file (.obj) with given vertices and lines connecting them sequentially.
    Args:
        file_path (str): The path to save the OBJ file.
        vertices (list of tuples): A list of (x, y, z) tuples representing the vertices.
    """
    with open(file_path, 'w') as obj_file:
        obj_file.write("# OBJ file\n")
        for vertex in vertices:
            x, y, z = vertex
            obj_file.write(f"v {x} {y} {z}\n")
        for i in range(len(vertices) - 1):
            obj_file.write(f"l {i + 1} {i + 2}\n")  # OBJ indices start at 1

def write_ply_with_vertices_and_edges(filename, list_curve_points):
    with open(filename, 'w') as ply_file:
        ply_file.write("ply\n")
        ply_file.write("format ascii 1.0\n")
        ply_file.write(f"element vertex {len(list_curve_points)}\n")
        ply_file.write("property float x\n")
        ply_file.write("property float y\n")
        ply_file.write("property float z\n")
        
        # Specify edges
        ply_file.write(f"element edge {len(list_curve_points) - 1}\n")
        ply_file.write("property int vertex1\n")
        ply_file.write("property int vertex2\n")
        ply_file.write("end_header\n")

        # Write vertices
        for point in list_curve_points:
            x, y, z = point
            ply_file.write(f"{x:.6f} {y:.6f} {z:.6f}\n")

        # Write edges
        for i in range(len(list_curve_points) - 1):
            ply_file.write(f"{i} {i + 1}\n")  # Ensure edges are written correctly

# visualize savety zones
import os
import numpy as np
def write_ply_with_vertices_and_faces_safety_zones(ply_file_path, zone, min_max_clearance):
    vertices = []
    faces = []
    
    min_z, max_z = min_max_clearance  # Unpack the min and max clearance values

    # Create vertices for the polygon at min_z and max_z
    num_vertices = len(zone)
    for (x, y) in zone:
        vertices.append((x, y, min_z))  # vertices at min_z
        vertices.append((x, y, max_z))  # vertices at max_z

    # Create side faces
    for i in range(num_vertices):
        next_i = (i + 1) % num_vertices
        # Connect min_z and max_z vertices
        faces.append(f"3 {i*2} {next_i*2} {next_i*2 + 1}")
        faces.append(f"3 {i*2} {next_i*2 + 1} {i*2 + 1}")

    # Create bottom and top faces using a fan triangulation method
    # Reference point for the fan (first vertex of the list)
    ref_min = 0
    ref_max = 1
    for i in range(1, num_vertices - 1):
        min_i = i * 2
        max_i = i * 2 + 1
        next_min = (i + 1) * 2
        next_max = (i + 1) * 2 + 1
        # Bottom plane triangle
        faces.append(f"3 {ref_min} {min_i} {next_min}")
        # Top plane triangle
        faces.append(f"3 {ref_max} {next_max} {max_i}")

    # Write to PLY file
    with open(ply_file_path, 'w') as file:
        file.write("ply\n")
        file.write("format ascii 1.0\n")
        file.write(f"element vertex {2 * num_vertices}\n")
        file.write("property float x\n")
        file.write("property float y\n")
        file.write("property float z\n")
        file.write(f"element face {4 * num_vertices + 2 * (num_vertices - 2)}\n")  # Adjusted face count
        file.write("property list uchar int vertex_index\n")
        file.write("end_header\n")
        for v in vertices:
            file.write(f"{v[0]} {v[1]} {v[2]}\n")
        for f in faces:
            file.write(f"{f}\n")



def append_tags_to_route(flight_route, tag):
    """Append a tag to each point in the flight route."""
    tagged_route = []
    for point in flight_route:
        if len(point) == 3:  # Assuming points are [x, y, z]
            tagged_point = point + [tag]  # Append the section tag
        else:
            tagged_point = point[:3] + [tag]  # Replace existing tag if there's more data
        tagged_route.append(tagged_point)
    return tagged_route

def flatten_points(route):
    flattened_route = []
    for point in route:
        # Ensure the point is not nested
        if isinstance(point[0], list):
            # Assume the point is a list of lists with only one inner list
            flattened_route.append(point[0])
        else:
            flattened_route.append(point)
    return flattened_route


def apply_transformation_to_route(route, dx=0, dy=0, dz=0):
    transformed_route = []
    for point in route:
        # Check and convert data types
        try:
            x = float(point[0])
            y = float(point[1])
            z = float(point[2])
        except ValueError as e:
            raise ValueError(f"Error converting point to float: {point} - {e}")
        except IndexError as e:
            raise IndexError(f"Missing coordinate data in point: {point} - {e}")

        # Apply transformations
        transformed_point = [x + dx, y + dy, z + dz]

        # Append additional data if present
        if len(point) > 3:
            transformed_point.extend(point[3:])

        transformed_route.append(transformed_point)

    return transformed_route
